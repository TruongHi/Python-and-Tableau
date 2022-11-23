# -*- coding: utf-8 -*-
"""
Created on Tue Oct 18 04:03:22 2022

@author: USER
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

excel_file = pd.read_csv('C:/Users/USER/.spyder-py3/Project 13 Supermarket sales/supermarket_sales - Sheet1.csv')
excel_file[['Gender', 'Product line', 'Total']]

excel_file.duplicated().sum()

excel_file.info()

report_table = excel_file.pivot_table(index='Gender',
                                      columns='Product line',
                                      values='Total',
                                      aggfunc='sum').round(0)

report_table.to_excel('report_2021.xlsx',
                      sheet_name='Report',
                      startrow=4)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# cell references (original spreadsheet) 
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# barchart
barchart = BarChart()

#locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row) #including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row) #not including headers

# adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)
#location chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 5 #choose the chart style
wb.save('report_2021.xlsx')

# Applying Excel formulas through Python

alphabet = list(string.ascii_uppercase)
excel_alphabet = alphabet[0:max_column] 
print(excel_alphabet)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']

# sum in columns B-G
for i in excel_alphabet:
    if i!='A':
        sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        sheet[f'{i}{max_row+1}'].style = 'Currency'
# adding total label
sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
wb.save('report_2021.xlsx')

# Formatting the report sheet

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
sheet['A1'] = 'Sales Report'
sheet['A2'] = '2021'
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)
wb.save('report_2021.xlsx')

# Automating the Report with a Python Function

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

def automate_excel(file_name):
    """The file name should have the following structure: sales_month.xlsx"""
    # read excel file
    excel_file = pd.read_excel(file_name)
    # make pivot table
    report_table = excel_file.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
    # splitting the month and extension from the file name
    month_and_extension = file_name.split('_')[1]
    # send the report table to excel file
    report_table.to_excel(f'report_{month_and_extension}', sheet_name='Report', startrow=4)
    # loading workbook and selecting sheet
    wb = load_workbook(f'report_{month_and_extension}')
    sheet = wb['Report']
    # cell references (original spreadsheet)
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    # adding a chart
    barchart = BarChart()
    data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row) #including headers
    categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row) #not including headers
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)
    sheet.add_chart(barchart, "B12") #location chart
    barchart.title = 'Sales by Product line'
    barchart.style = 2 #choose the chart style
    # applying formulas
    # first create alphabet list as references for cells
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column] #note: Python lists start on 0 -> A=0, B=1, C=2. #note2 the [a:b] takes b-a elements
    # sum in columns B-G
    for i in excel_alphabet:
        if i!='A':
            sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
            sheet[f'{i}{max_row+1}'].style = 'Currency'
    sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
    # getting month name
    month_name = month_and_extension.split('.')[0]
    # formatting the report
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = month_name.title()
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)
    wb.save(f'report_{month_and_extension}')
    return 

# Applying the function to multiple Excel files
# read excel files
excel_file_1 = pd.read_excel('sales_january.xlsx')
excel_file_2 = pd.read_excel('sales_february.xlsx')
excel_file_3 = pd.read_excel('sales_march.xlsx')

# concatenate files
new_file = pd.concat([excel_file_1,
                      excel_file_2,
                      excel_file_3], ignore_index=True)

# export file
new_file.to_excel('sales_2021.xlsx')
# apply function
automate_excel('sales_2021.xlsx')


